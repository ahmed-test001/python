import pandas as pd
from openpyxl import load_workbook


class ExcelUtil():
    def __init__(self, tc_name, file_name="../Utility_TestData_Excel/2021/w_08/W_08_TerraQ_PO_T2_version_2.xlsx"):
    # def __init__(self, tc_name, file_name="C:/Users/a.ferdous.CORP/PycharmProjects/com.CheckProofing/Utility_TestData_Excel/2021/w_04/W_04_Palette_PO_T3.xlsx"):
        # self.tc_name = tc_name.replace("test_","")
        self.tc_name = tc_name.replace("", "")
        self.file_name = file_name
        self.df = pd.read_excel(file_name)
        self.workbook = load_workbook(filename=self.file_name)
        # self.sheet = self.workbook['Sheet1']

    def read_by_tc(self, tc_name):
        tc_df = self.df[self.df["TC_Name"] == tc_name]
        try:
            tc_dict = tc_df.to_dict('records')[0]
        except IndexError:
            raise RuntimeError("Test Case {} not found in excel sheet. Please check.".format(tc_name))

        return tc_dict

    def read_by_column(self, i, col):
        self.sheet = self.workbook[i]
        tc_dict = self.read_by_tc(self.tc_name)
        value = tc_dict.get(col)
        return value

    def read_from_excel(self,i ,row, col):
            self.sheet = self.workbook[i]
            value = self.sheet.cell(row=row, column=col).value
            return value

    def write_to_excel(self, row, col, value):
        self.sheet.cell(row=row+1, column=col).value = value
        self.workbook.save(self.file_name)

        # Refresh dataframe
        self.df = pd.read_excel(self.file_name)
        print("Value {} written successfully to row {} col {}".format(value, row, col))


# For Testing
if __name__ == "__main__":
    obj = ExcelUtil("get_pay_later_icon_LandingPage_URL_validation")
    # obj.write_to_excel(5,5,"test")
    # obj.read_by_tc("a1_subjectLine_text_validation")






















# import openpyxl
# import xlrd as xlrd


# class ExcelReaderUtil:
#     @staticmethod
#     def getTestData(test_case_name):
#         Dict = {}
#         book = openpyxl.load_workbook("../worksheet.xlsx")
#         sheet = book.active
#         for i in range(1, sheet.max_row + 1):  # to get rows
#             if sheet.cell(row=i, column=1).value == test_case_name:
#                 for j in range(2, sheet.max_column + 1):  # to get columns
#                     Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
#         return [Dict]

    # def getTestData(self):
    #     testdata=[{sl,ph}]


# Dict = {}
# book = openpyxl.load_workbook("../worksheet.xlsx")
# sheet = book.active
# for i in range(1, sheet.max_row + 1):  # to get rows
#     if sheet.cell(row=i, column=1).value == "testexl":
#         for j in range(2, sheet.max_column + 1):  # to get columns
#             Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
# print(Dict)



    # workbook = xlrd.open_workbook("../worksheet.xlsx")
    # sheet = workbook.sheet_by_name("Sheet1")
    #
    # # Get number of rows with data in excel sheet
    # rowcount = sheet.nrows
    # # Get number of columns with data in each row. Returns highest number
    # colcount = sheet.ncols
    # print(rowcount)
    # print(colcount)
    #
    # result_data = []
    # for curr_row in range(1, rowcount, 1):
    #     row_data = []
    #
    #     for curr_col in range(1, colcount - 1, 1):
    #         # Read the data in the current cell
    #         data = sheet.cell_value(curr_row, curr_col)
    #         print(data)
    #         row_data.append(data)
    #
    #     result_data.append(row_data)